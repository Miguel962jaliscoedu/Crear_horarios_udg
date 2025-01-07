import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta
import os

# Función para transformar el archivo Excel subido
def transform_excel(input_df):
    columns_mapping = {
        "Unnamed: 1": "NRC",
        "Unnamed: 2": "Clave",
        "Unnamed: 3": "Materia",
        "Unnamed: 4": "Sec",
        "Unnamed: 5": "CR",
        "Unnamed: 6": "CUP",
        "Unnamed: 7": "DIS",
        "Unnamed: 8": "Ses",
        "Unnamed: 9": "Hora",
        "Unnamed: 10": "Días",
        "Unnamed: 11": "Edif",
        "Unnamed: 12": "Aula",
        "Unnamed: 15": "Profesor"
    }
    input_df = input_df.rename(columns=columns_mapping)
    output_columns = [
        "NRC", "Clave", "Materia", "Sec", "CR", "CUP", "DIS",
        "Ses", "Hora", "Días", "Edif", "Aula", "Profesor"
    ]
    output_df = input_df[output_columns]
    output_df = output_df.dropna(how='all').reset_index(drop=True)
    return output_df

# Modificar días a nombres válidos
days_mapping = {
    "L": "Lunes", "M": "Martes", "I": "Miércoles", "J": "Jueves", "V": "Viernes",
    "lunes": "Lunes", "martes": "Martes", "miércoles": "Miércoles", "jueves": "Jueves", "viernes": "Viernes",
    "LUNES": "Lunes", "MARTES": "Martes", "MIÉRCOLES": "Miércoles", "JUEVES": "Jueves", "VIERNES": "Viernes",
}

# Nueva función para limpiar los días
def clean_days(value):
    if isinstance(value, str):
        # Descomponer en caracteres si es necesario
        possible_days = value.strip().split()  # Separar por espacios si es necesario
        cleaned_days = [days_mapping.get(day, None) for day in possible_days]
        return [day for day in cleaned_days if day is not None]
    return []

# Función para procesar y transformar los datos
def process_data(first_sheet):
    first_sheet = first_sheet.iloc[1:]
    columns_to_extract = ["NRC", "Materia", "Sec", "Ses", "Hora", "Días", "Edif", "Aula", "Profesor"]
    first_sheet = first_sheet[columns_to_extract]
    first_sheet.columns = ["NRC", "Materia", "Sección", "Sesión", "Hora", "Días", "Edificio", "Aula", "Profesor"]
    first_sheet.loc[:, "Materia"] = first_sheet["Materia"].ffill()
    first_sheet.loc[:, "NRC"] = first_sheet["NRC"].ffill()
    first_sheet.loc[:, "Sección"] = first_sheet["Sección"].ffill()
    first_sheet.loc[:, "Profesor"] = first_sheet["Profesor"].ffill()
    first_sheet = first_sheet.dropna(subset=["Sesión", "Hora", "Días"])

    # Aplicar limpieza en los días
    first_sheet["Días"] = first_sheet["Días"].apply(clean_days)

    # Depuración: Verificar días únicos después de limpiar
    unique_days = first_sheet["Días"].explode().unique()
    st.write("Días únicos después de limpiar:", unique_days)

    def parse_time_range(time_string):
        try:
            start, end = time_string.split("-")
            start_dt = pd.to_datetime(start.strip(), format="%H%M")
            end_dt = pd.to_datetime(end.strip(), format="%H%M")
            start_12h = start_dt.strftime("%I:%M %p")
            end_12h = end_dt.strftime("%I:%M %p")
            return f"{start_12h} - {end_12h}"
        except Exception:
            return time_string

    first_sheet["Hora"] = first_sheet["Hora"].apply(parse_time_range)
    expanded_data = first_sheet.explode("Días").reset_index(drop=True)
    return expanded_data

# Función para filtrar los datos según el NRC
def filter_by_nrc(data):
    st.write("Datos originales:")
    st.dataframe(data)

    unique_nrcs = data[["NRC", "Materia", "Profesor"]].drop_duplicates()
    st.write("Lista única de NRC disponibles:")
    st.dataframe(unique_nrcs)

    selected_nrcs = st.multiselect("Selecciona los NRC de las materias que deseas incluir:", options=unique_nrcs["NRC"].astype(str).tolist())

    # Mostrar selección del usuario
    st.write("NRC seleccionados por el usuario:", selected_nrcs)

    # Filtrar los datos por NRC
    data["NRC"] = data["NRC"].astype(str)  # Asegurarse de que sean cadenas
    filtered_data = data[data["NRC"].isin(selected_nrcs)]

    if not filtered_data.empty:
        st.success(f"Se encontraron {len(filtered_data)} registros para los NRC seleccionados.")
    else:
        st.warning("No hay datos disponibles para los NRC seleccionados. Revisa los valores seleccionados.")

    # Mostrar datos filtrados para depuración
    st.write("Datos filtrados:")
    st.dataframe(filtered_data)

    return filtered_data


# Función para crear una hoja de horario
def create_schedule_sheet(expanded_data):
    hours_list = [
        "07:00 AM - 07:59 AM", "08:00 AM - 08:59 AM", "09:00 AM - 09:59 AM", "10:00 AM - 10:59 AM",
        "11:00 AM - 11:59 AM", "12:00 PM - 12:59 PM", "01:00 PM - 01:59 PM", "02:00 PM - 02:59 PM",
        "03:00 PM - 03:59 PM", "04:00 PM - 04:59 PM", "05:00 PM - 05:59 PM", "06:00 PM - 06:59 PM",
        "07:00 PM - 07:59 PM"
    ]
    days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    schedule = pd.DataFrame(columns=["Hora"] + days)
    schedule["Hora"] = hours_list

    for index, row in expanded_data.iterrows():
        for hour_range in hours_list:
            start_hour, end_hour = [pd.to_datetime(hr, format="%I:%M %p") for hr in hour_range.split(" - ")]
            class_start, class_end = [pd.to_datetime(hr, format="%I:%M %p") for hr in row["Hora"].split(" - ")]
            if start_hour < class_end and class_start < end_hour:
                day_col = row["Días"]
                content = f"{row['Materia']} ({row['Aula']})\n{row['Profesor']}"
                if pd.notna(schedule.loc[schedule["Hora"] == hour_range, day_col].values[0]):
                    schedule.loc[schedule["Hora"] == hour_range, day_col] += "\n" + content
                else:
                    schedule.loc[schedule["Hora"] == hour_range, day_col] = content
    return schedule

# Función principal
def main():
    st.title("Generador de Horarios a partir de Excel")

    uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
    if uploaded_file is not None:
        input_df = pd.ExcelFile(uploaded_file).parse(0)
        st.write("Vista previa del archivo original:")
        st.dataframe(input_df.head())

        output_df = transform_excel(input_df)
        st.write("Archivo transformado:")
        st.dataframe(output_df)

        expanded_data = process_data(output_df)
        st.write("Datos procesados:")
        st.dataframe(expanded_data)

        filtered_data = filter_by_nrc(expanded_data)
        if not filtered_data.empty:
            if st.button("Generar horario"):
                schedule_df = create_schedule_sheet(filtered_data)
                st.write("Horario generado:")
                st.dataframe(schedule_df)

                # Descargar el horario como archivo Excel
                output_path = "horario_generado.xlsx"
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    schedule_df.to_excel(writer, index=False, sheet_name="Horario")
                with open(output_path, "rb") as f:
                    st.download_button(
                        label="Descargar horario generado",
                        data=f,
                        file_name="horario_generado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == "__main__":
    main()
